"""
eHistorian Gateway — Unified Server (API + Web UI)
"""

import json
import os
import subprocess
from pathlib import Path
from datetime import datetime
from flask import Flask, render_template, jsonify, request
from flask_cors import CORS
from flask_cors import logging

# ── TEST ONLY: CAN BE EASILY DELETED ──
try:
    from tests.test_component import attach_test_routes, is_outage_simulated, is_invalid_config_simulated
except ImportError:
    def is_outage_simulated(): return False
    def is_invalid_config_simulated(): return False
# ── END TEST ONLY ──

# ui/ je ve stejne slozce jako tento soubor
BASE_DIR = Path(__file__).parent
UI_DIR   = BASE_DIR / 'ui'

app = Flask(
    __name__,
    template_folder=str(UI_DIR),   # index.html
    static_folder=str(UI_DIR),     # styles.css, app.js
    static_url_path=''             # /styles.css misto /static/styles.css
)
CORS(app)

LOGS_DIR    = BASE_DIR / 'logs'
LOGS_DIR.mkdir(exist_ok=True)

FAKEGEN_CONFIG_FILE = BASE_DIR / 'simulator' / 'fakegen_config.json'



# ── Process handles ─────────────────────────────────────
processes = {
    'gateway': None,
    'fakegen': None,
    'opcuaserver': None
}

# ── Activity log buffer ─────────────────────────────────
from collections import deque
_activity_log: deque = deque(maxlen=500)
_activity_last_id: int = 0

def log_activity(msg: str, kind: str = 'info') -> None:
    """Appends a timestamped entry to the in-memory activity ring-buffer."""
    global _activity_last_id
    _activity_last_id += 1
    _activity_log.append({
        'id':        _activity_last_id,
        'timestamp': datetime.utcnow().isoformat() + 'Z',
        'kind':      kind,   # info | ok | warn | error | process
        'msg':       msg
    })

# ── Config helpers ───────────────────────────────────────
CONFIGS_DIR = BASE_DIR / 'configs'
CONFIGS_DIR.mkdir(exist_ok=True)
CONFIG_FILE = CONFIGS_DIR / 'default.json'

def load_config() -> dict:
    if CONFIG_FILE.exists():
        try:
            with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            pass
    return {
        "gatewayId": "line-01-secret",
        "apiUrl":    "http://localhost:5000",
        "opcua":     [],
        "sql":       [],
        "offlineBufferMaxBytes": 10485760
    }

def save_config(config: dict) -> None:
    with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
        json.dump(config, f, indent=2, ensure_ascii=False)

# ── Process helpers ──────────────────────────────────────
def start_process(name: str, cmd: str) -> bool:
    try:
        processes[name] = subprocess.Popen(cmd, shell=True)
        return True
    except Exception as e:
        print(f"[ERROR] Starting {name}: {e}")
        return False

def get_psutil():
    try:
        import psutil
        return psutil
    except ImportError:
        import subprocess, sys
        try:
            print("[INFO] Instaluji psutil pro sledování procesů...")
            subprocess.check_call([sys.executable, "-m", "pip", "install", "psutil", "-q"])
            import psutil
            return psutil
        except Exception as e:
            print(f"[ERROR] Nelze nainstalovat psutil: {e}")
            return None

def stop_process(name: str) -> bool:
    killed = False
    try:
        psutil = get_psutil()
        if psutil:
            for proc in psutil.process_iter(['name', 'cmdline']):
                try:
                    pinfo = proc.info
                    if pinfo.get('name') and 'python' in pinfo['name'].lower():
                        cmdline = pinfo.get('cmdline') or []
                        cmd_str = ' '.join(cmdline).lower()
                        
                        is_target = False
                        if name == 'gateway' and 'ehistorian_gateway.main' in cmd_str:
                            is_target = True
                        elif name == 'fakegen' and 'fake_data_generator.py' in cmd_str:
                            is_target = True
                        elif name == 'opcuaserver' and 'fake_opcua_server.py' in cmd_str:
                            is_target = True
                            
                        if is_target:
                            proc.terminate()
                            try:
                                proc.wait(timeout=3)
                            except psutil.TimeoutExpired:
                                proc.kill()
                            killed = True
                except (psutil.NoSuchProcess, psutil.AccessDenied, psutil.ZombieProcess):
                    pass
    except Exception as e:
        print(f"[ERROR] Stopping {name}: {e}")
        
    if killed:
        return True
        
    proc = processes.get(name)
    if proc is not None:
        try:
            proc.terminate()
        except:
            pass
        processes[name] = None
    return True

def is_running(name: str) -> bool:
    try:
        psutil = get_psutil()
        if psutil:
            for proc in psutil.process_iter(['name', 'cmdline']):
                try:
                    pinfo = proc.info
                    if pinfo.get('name') and 'python' in pinfo['name'].lower():
                        cmdline = pinfo.get('cmdline') or []
                        cmd_str = ' '.join(cmdline).lower()
                        
                        if name == 'gateway' and 'ehistorian_gateway.main' in cmd_str:
                            return True
                        if name == 'fakegen' and 'fake_data_generator.py' in cmd_str:
                            return True
                        if name == 'opcuaserver' and 'fake_opcua_server.py' in cmd_str:
                            return True
                except (psutil.NoSuchProcess, psutil.AccessDenied, psutil.ZombieProcess):
                    pass
    except Exception as e:
        print(f"[ERROR] is_running: {e}")
        
    proc = processes.get(name)
    return proc is not None and proc.poll() is None

# ── Routes: UI ────────────────────────────────────────────────
@app.route('/')
def index():
    return render_template('index.html')

# Activity feed
@app.route('/api/activity', methods=['GET'])
def get_activity():
    since_id = int(request.args.get('since', 0))
    entries = [e for e in _activity_log if e['id'] > since_id]
    return jsonify(entries)



# Config (UI)
@app.route('/api/config', methods=['GET', 'POST'])
def ui_config():
    if request.method == 'GET':
        return jsonify(load_config())
    config_data = request.get_json()
    save_config(config_data)
    sql_count = len(config_data.get('sql', []))
    opc_count = len(config_data.get('opcua', []))
    log_activity(f"Config saved — SQL: {sql_count} | OPC UA: {opc_count}", 'ok')
    return jsonify({'status': 'saved'})

# Fix SQL drivers
@app.route('/api/config/fix-drivers', methods=['POST'])
def fix_drivers():
    try:
        import pyodbc
        import re
    except ImportError:
        return jsonify({'status': 'error', 'message': 'Missing pyodbc library'}), 500

    drivers = pyodbc.drivers()
    if not drivers:
        return jsonify({'status': 'error', 'message': 'Žádné ODBC ovladače nebyly nalezeny.'}), 404

    best_driver = None
    priorities = ["ODBC Driver 18 for SQL Server", "ODBC Driver 17 for SQL Server", "SQL Server"]
    for p in priorities:
        if p in drivers:
            best_driver = p
            break
            
    if not best_driver:
        return jsonify({'status': 'error', 'message': f'Nenalezen žádný kompatibilní ovladač. Dostupné: {drivers}'}), 404

    config = load_config()
    changed = False
    
    for src in config.get('sql', []):
        conn_str = src.get('connectionString', '')
        if conn_str:
            new_conn_str = re.sub(r'Driver=\{[^}]+\}', f'Driver={{{best_driver}}}', conn_str, flags=re.IGNORECASE)
            
            if "ODBC Driver 18" in best_driver:
                if "TrustServerCertificate=yes" not in new_conn_str and "TrustServerCertificate=Yes" not in new_conn_str:
                    new_conn_str = re.sub(r'TrustServerCertificate=no;?', '', new_conn_str, flags=re.IGNORECASE)
                    if not new_conn_str.endswith(';'):
                        new_conn_str += ';'
                    new_conn_str += 'TrustServerCertificate=yes;'
                    
            if new_conn_str != conn_str:
                src['connectionString'] = new_conn_str
                changed = True

    if changed:
        save_config(config)
        return jsonify({'status': 'ok', 'message': f'Ovladač nastaven na {best_driver}'})
    else:
        return jsonify({'status': 'ok', 'message': 'Konfigurace je aktuální.'})

# Status
@app.route('/api/status', methods=['GET'])
def status():
    return jsonify({
        'gateway': is_running('gateway'),
        'fakegen': is_running('fakegen'),
        'opcuaserver': is_running('opcuaserver')
    })

# Device IP
@app.route('/api/ip', methods=['GET'])
def get_ip():
    try:
        import urllib.request
        req = urllib.request.Request('https://api.ipify.org')
        with urllib.request.urlopen(req, timeout=3) as response:
            ip = response.read().decode('utf8')
        return jsonify({'ip': ip})
    except Exception as e:
        return jsonify({'ip': 'unknown', 'error': str(e)})

# Fakegen config
@app.route('/api/fakegen/config', methods=['GET', 'POST'])
def fakegen_config():
    default = {
        "connectionString": (
            "Driver={ODBC Driver 18 for SQL Server};"
            "Server=localhost;"
            "Database=eFactory;"
            "UID=sa;"
            "PWD=YourStrong!Passw0rd;"
            "TrustServerCertificate=yes;"
        ),
        "table": "CurrentValues",
        "tagColumn": "TagName",
        "valueColumn": "Value",
        "timestampColumn": "UpdatedAt",
        "intervalSeconds": 10,
        "sensors": {
            "Temperature": {"base": 22.0, "noise": 1.5},
            "Pressure": {"base": 1013.25, "noise": 10.0},
            "Flow": {"base": 50.0, "noise": 5.0},
            "Level": {"base": 75.0, "noise": 2.0}
        }
    }
    if request.method == 'GET':
        if FAKEGEN_CONFIG_FILE.exists():
            with open(FAKEGEN_CONFIG_FILE, 'r', encoding='utf-8') as f:
                return jsonify(json.load(f))
        return jsonify(default)
    data = request.get_json() or {}
    with open(FAKEGEN_CONFIG_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    log_activity(f"Fakegen config saved — db: {data.get('table','?')} | interval: {data.get('intervalSeconds','?')}s | sensors: {len(data.get('sensors',{}))}", 'ok')
    return jsonify({'status': 'saved'})

# Process control
@app.route('/api/processes/<name>', methods=['POST'])
def manage_process(name: str):
    if name not in processes:
        return jsonify({'status': 'error', 'message': 'Unknown process'}), 400

    action = (request.get_json() or {}).get('action')

    if action == 'start':
        if name == 'gateway':
            env = os.environ.copy()
            env["EHG_BOOTSTRAP_CONFIG"] = str(CONFIG_FILE)
            env["EHG_NO_PAUSE"] = "1"
            try:
                processes[name] = subprocess.Popen(
                    'start cmd /c "set EHG_NO_PAUSE=1&&run_gateway.bat"',
                    shell=True,
                    cwd=str(BASE_DIR),
                    env=env
                )
                success = True
            except Exception as e:
                print(f"[ERROR] Starting {name}: {e}")
                success = False

        elif name == 'fakegen':
            env = os.environ.copy()
            # Ensure config file exists before starting
            if not FAKEGEN_CONFIG_FILE.exists():
                fakegen_config()  # Creates default if missing
            env["FAKEGEN_CONFIG"] = str(FAKEGEN_CONFIG_FILE)
            env["EHG_NO_PAUSE"] = "1"
            try:
                processes[name] = subprocess.Popen(
                    'start cmd /c "set EHG_NO_PAUSE=1&&run_fake_data_generator.bat"',
                    shell=True,
                    cwd=str(BASE_DIR),
                    env=env
                )
                success = True
                log_activity(f"Fakegen started in separate CMD", 'process')
            except Exception as e:
                print(f"[ERROR] Starting {name}: {e}")
                log_activity(f"Fakegen start failed: {e}", 'error')
                success = False

        elif name == 'opcuaserver':
            env = os.environ.copy()
            env["EHG_NO_PAUSE"] = "1"
            try:
                processes[name] = subprocess.Popen(
                    'start cmd /c "set EHG_NO_PAUSE=1&&run_fake_opcua_server.bat"',
                    shell=True,
                    cwd=str(BASE_DIR),
                    env=env
                )
                success = True
                log_activity(f"OPC UA Simulator started in separate CMD", 'process')
            except Exception as e:
                print(f"[ERROR] Starting {name}: {e}")
                log_activity(f"OPC UA Simulator start failed: {e}", 'error')
                success = False

        else:
            return jsonify({'status': 'error', 'message': 'No command defined'}), 400

        return jsonify({'status': 'started' if success else 'error', 'process': name})

    elif action == 'stop':
        stop_process(name)
        log_activity(f"{name.capitalize()} stopped", 'warn')
        return jsonify({'status': 'stopped', 'process': name})

    return jsonify({'status': 'error', 'message': 'Unknown action'}), 400

# Test event
@app.route('/api/test', methods=['POST'])
def run_test():
    try:
        import requests as req

        url  = "http://localhost:5000/api/ehistorian/gateway/ingest"
        data = {
            "gatewayId": load_config().get('gatewayId', 'test-gateway'),
            "events": [
                {
                    "gatewayId": "test-gateway",
                    "assetId":   101,
                    "source":    "test",
                    "sourceId":  "test-0:asset-101",
                    "tag":       "Temperature",
                    "value":     20 + (hash(str(datetime.now())) % 10),
                    "timestamp": datetime.utcnow().isoformat() + "Z",
                    "quality":   "Good"
                }
            ]
        }

        response = req.post(url, json=data, timeout=5)
        return jsonify({'status': 'ok', 'response': response.json()})

    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

# Logs — list
@app.route('/api/logs', methods=['GET'])
def list_logs():
    log_files = sorted(LOGS_DIR.glob('*.json'), reverse=True)
    result = []
    for f in log_files:
        try:
            size = f.stat().st_size
            result.append({
                'name':          f.name,
                'timestamp':     datetime.fromtimestamp(f.stat().st_mtime).isoformat(),
                'size':          size,
                'size_readable': f"{size / 1024:.1f} KB" if size > 1024 else f"{size} B"
            })
        except Exception:
            pass
    return jsonify(result)

# Logs — get single
@app.route('/api/log/<filename>', methods=['GET'])
def get_log(filename: str):
    path = LOGS_DIR / filename
    if not path.exists() or path.suffix != '.json':
        return jsonify({'error': 'Not found'}), 404
    try:
        with open(path, 'r', encoding='utf-8') as f:
            return jsonify(json.load(f))
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Logs — delete single
@app.route('/api/log/<filename>', methods=['DELETE'])
def delete_log(filename: str):
    path = LOGS_DIR / filename
    if not path.exists():
        return jsonify({'error': 'Not found'}), 404
    try:
        path.unlink()
        return jsonify({'status': 'deleted'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Logs — clear all
@app.route('/api/logs/clear', methods=['POST'])
def clear_logs():
    try:
        deleted = 0
        for f in LOGS_DIR.glob('*.json'):
            f.unlink()
            deleted += 1
        return jsonify({'status': 'cleared', 'deleted': deleted})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Logs — export to Excel
@app.route('/api/logs/export/excel', methods=['GET'])
def export_logs_excel():
    try:
        import pandas as pd
        from openpyxl.chart import LineChart, Reference
        from openpyxl.chart.axis import DateAxis
    except ImportError:
        import subprocess, sys
        try:
            subprocess.check_call([sys.executable, "-m", "pip", "install", "pandas", "openpyxl", "-q"])
            import pandas as pd
            from openpyxl.chart import LineChart, Reference
            from openpyxl.chart.axis import DateAxis
        except Exception as e:
            return jsonify({'error': f"Failed to install pandas/openpyxl: {e}"}), 500

    try:
        from io import BytesIO
        from flask import send_file
        
        all_events = []
        for f in LOGS_DIR.glob('*.json'):
            try:
                with open(f, 'r', encoding='utf-8') as json_file:
                    data = json.load(json_file)
                    gw_id = data.get('gatewayId', '')
                    for ev in data.get('events', []):
                        ev['gatewayId'] = gw_id
                        all_events.append(ev)
            except Exception:
                pass
                
        if not all_events:
            return jsonify({'error': 'No logs available to export'}), 404

        df = pd.DataFrame(all_events)
        
        if 'timestamp' in df.columns:
            df['timestamp'] = pd.to_datetime(df['timestamp'], errors='coerce', utc=True)
            df['timestamp'] = df['timestamp'].dt.tz_localize(None)  # strip tz for Excel
            df = df.sort_values(by='timestamp')

        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Raw Data', index=False)

            # ── Format Raw Data sheet ──────────────────────────────
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            ws = writer.sheets['Raw Data']
            ws.freeze_panes = 'A2'

            col_widths = {}
            # Set column widths and number format
            for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                for cell in row:
                    # Number format for value column
                    if cell.column_letter == 'F' or (cell.value is not None and isinstance(cell.value, float)):
                        cell.number_format = '#,##0.00'
                    # Track max content width
                    col_letter = cell.column_letter
                    val_len = len(str(cell.value)) if cell.value is not None else 0
                    col_widths[col_letter] = max(col_widths.get(col_letter, 8), val_len)

            # Apply header widths too
            for i, cell in enumerate(ws[1], 1):
                col_letter = get_column_letter(i)
                header_len = len(str(cell.value)) if cell.value else 8
                col_widths[col_letter] = max(col_widths.get(col_letter, 8), header_len)

            for col_letter, width in col_widths.items():
                ws.column_dimensions[col_letter].width = min(width + 4, 50)

            ws.row_dimensions[1].height = 22

            numeric_df = df.dropna(subset=['value']).copy()
            numeric_df = numeric_df[pd.to_numeric(numeric_df['value'], errors='coerce').notnull()]
            numeric_df['value'] = pd.to_numeric(numeric_df['value'])

            if 'tag' in numeric_df.columns:
                tags = numeric_df['tag'].unique()
                valid_tags = [t for t in tags if len(numeric_df[numeric_df['tag'] == t]) >= 2]

                if valid_tags:
                    workbook = writer.book
                    from openpyxl.styles import Font as XFont, PatternFill as XFill, Alignment as XAlign
                    from openpyxl.utils import get_column_letter

                    data_ws = workbook.create_sheet('Charts & Data')

                    # Line colors for up to 8 tags
                    LINE_COLORS = ['4472C4', 'ED7D31', 'A9D18E', 'FF0000',
                                   '7030A0', '00B0F0', 'FFC000', '70AD47']

                    tag_col_map = {}
                    max_data_rows = 0

                    for t_idx, tag in enumerate(valid_tags):
                        tc = t_idx * 2 + 1   # 1-based: 1,3,5,7...
                        vc = tc + 1
                        tag_col_map[tag] = (tc, vc)

                        # Column headers
                        th = data_ws.cell(row=1, column=tc, value=f"{tag} Time")
                        vh = data_ws.cell(row=1, column=vc, value=f"{tag} Value")

                        # Data rows
                        tag_df = numeric_df[numeric_df['tag'] == tag].copy()
                        if len(tag_df) > max_data_rows:
                            max_data_rows = len(tag_df)
                            
                        for row_i, (_, r) in enumerate(tag_df.iterrows(), start=2):
                            ts = r['timestamp']
                            label = ts.strftime('%d.%m %H:%M:%S') if pd.notnull(ts) and hasattr(ts, 'strftime') else str(ts)
                            data_ws.cell(row=row_i, column=tc, value=label)
                            v_cell = data_ws.cell(row=row_i, column=vc, value=float(r['value']))
                            v_cell.number_format = '#,##0.00'

                        # Column widths
                        data_ws.column_dimensions[get_column_letter(tc)].width = 16
                        data_ws.column_dimensions[get_column_letter(vc)].width = 12

                    data_ws.freeze_panes = 'A2'

                    # ── Add Charts Below Data ──
                    # Place charts starting after max_data_rows, 2 charts per row
                    ROWS_PER_CHART = 22
                    COLS_PER_CHART = 9
                    start_chart_row = max_data_rows + 4

                    for c_idx, tag in enumerate(valid_tags):
                        tag_df = numeric_df[numeric_df['tag'] == tag]
                        n_rows = len(tag_df)
                        tc, vc = tag_col_map[tag]

                        chart = LineChart()
                        chart.title  = f"Trend — {tag}"
                        chart.style  = 2
                        chart.width  = 18
                        chart.height = 11
                        chart.y_axis.title = tag
                        chart.x_axis.title = 'Time'
                        chart.y_axis.numFmt = '#,##0.00'
                        chart.legend        = None

                        values = Reference(data_ws, min_col=vc, min_row=1, max_row=n_rows + 1)
                        cats   = Reference(data_ws, min_col=tc, min_row=2, max_row=n_rows + 1)

                        chart.add_data(values, titles_from_data=True)
                        chart.set_categories(cats)

                        s = chart.series[0]
                        color = LINE_COLORS[c_idx % len(LINE_COLORS)]
                        s.graphicalProperties.line.solidFill   = color
                        s.graphicalProperties.line.width        = 18000
                        s.marker.symbol = 'none'

                        grid_row = c_idx // 2
                        grid_col = c_idx %  2
                        anchor_row = start_chart_row + (grid_row * ROWS_PER_CHART)
                        anchor_col = grid_col * COLS_PER_CHART + 1
                        
                        data_ws.add_chart(chart, f"{get_column_letter(anchor_col)}{anchor_row}")


        output.seek(0)
        
        filename = f"eHistorian_Export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(
            output, 
            as_attachment=True, 
            download_name=filename, 
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ── TEST ONLY: CAN BE EASILY DELETED ──
    if request.method == 'POST':
        data = request.get_json() or {}
        GATEWAY_BUFFER_STATUS['bytesSize'] = data.get('bytesSize', 0)
        GATEWAY_BUFFER_STATUS['pendingCount'] = data.get('pendingCount', 0)
        GATEWAY_BUFFER_STATUS['maxBytes'] = data.get('maxBytes', GATEWAY_BUFFER_STATUS.get('maxBytes', 5000))
        return jsonify({'status': 'ok'})
    return jsonify(GATEWAY_BUFFER_STATUS)
# ── END TEST ONLY ──


# ── Routes: Gateway API ──────────────────────────────────
@app.route('/health', methods=['GET'])
def health():
    return jsonify({'status': 'ok'})

@app.route('/api/ehistorian/gateway/config/<gateway_id>', methods=['GET'])
def get_gateway_config(gateway_id):
    if is_invalid_config_simulated():
        from flask import Response
        print(f'[CONFIG] Gateway {gateway_id} requested config (Returning SIMULATED INVALID JSON)')
        return Response("{ toto_je_invalidni_json, chyba syntaxe: }", status=200, mimetype='application/json')
        
    config = load_config()
    sql_count = len(config.get('sql', []))
    opc_count = len(config.get('opcua', []))
    log_activity(f"GET config — GW: {gateway_id} | SQL: {sql_count} | OPC UA: {opc_count}", 'info')
    print(f'[CONFIG] Gateway {gateway_id} requested config (Returning {opc_count} OPC UA, {sql_count} SQL)')
    return jsonify(config)

@app.route('/api/ehistorian/gateway/ingest', methods=['POST'])
def ingest():
    if is_outage_simulated():
        print("[OUTAGE SIMULATION] Rejecting ingest request from gateway")
        return jsonify({'status': 'Rejected', 'error': 'Simulated Outage'}), 503

    data = request.get_json()
    gateway_id = data.get('gatewayId', 'unknown')
    events = data.get('events', [])

    log_file = LOGS_DIR / f'ingest_{datetime.now().strftime("%Y%m%d_%H%M%S_%f")}.json'

    with open(log_file, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2, ensure_ascii=False)

    # Log summary per tag
    summary = []
    for ev in events:
        val = ev.get('value','')
        if isinstance(val, float):
            val = f"{val:.2f}"
        summary.append(f"{ev.get('tag','')}={val}")

    if summary:
        log_activity(
            f"INGEST — Database Update: {gateway_id} | Inserted: {', '.join(summary)}",
            'ok'
        )

    print(f'[INGEST] {len(events)} events from {gateway_id} -> {log_file.name}')
    
    return jsonify({
        'gatewayId': gateway_id,
        'acceptedCount': len(events),
        'rejectedCount': 0,
        'status': 'Accepted'
    })


# ── TEST ONLY: CAN BE EASILY DELETED ──
try:
    attach_test_routes(app, load_config, BASE_DIR)
except NameError:
    pass
# ── END TEST ONLY ──

# ── Entry point ──────────────────────────────────────────
if __name__ == '__main__':
    print()
    print("=" * 50)
    print("  eHistorian Server & Control Panel")
    print("=" * 50)
    print(f"\n  Configs: {CONFIGS_DIR}")
    print(f"  Logs   : {LOGS_DIR}")
    print(f"  UI     : {UI_DIR}")
    print(f"\n  http://localhost:5000\n")
    app.run(host='0.0.0.0', port=5000, debug=False)
